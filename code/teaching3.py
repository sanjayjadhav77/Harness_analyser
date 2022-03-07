import global_var
import global_test_var as GV
from global_files import*
import teaching_3
from  Sql_db import *
import operator
import openpyxl
from HDM import *
from CuttingChart_import import*
class MyDelegate(QtGui.QItemDelegate):

    def createEditor(self, parent, option, index):
        editor = QtGui.QSpinBox(parent)
        editor.setMaximum(9999)
        return editor
class teaching_page3(QtGui.QMainWindow,teaching_3.Ui_MainWindow):   # class of contactus page

    def __init__(self):
        super(teaching_page3,self).__init__()
        self.setupUi(self)
        delegate = MyDelegate()                    # create delegate
        self.hrn_table.setItemDelegate(delegate)  # set delegate
        self.hrn_table.setEnabled(False) 
        self.pushButton_6.clicked.connect(self.save_hrn)
        self.pushButton_4.clicked.connect(self.edit_txt_data)
        self.pushButton.clicked.connect(self.first_stage_learn)
##        self.pushButton_5.clicked.connect(self.close_window)
        self.pushButton_7.clicked.connect(self.state_changed)
        self.comboBox_2.currentIndexChanged.connect(self.read_hrn_file_DB)
        
    def popupmeassage(self):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(self.quit_msg )
        msgBox.setWindowTitle("Message")
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)

        returnValue = msgBox.exec()
        if returnValue == QMessageBox.Ok:
            msgBox.close() 
        if returnValue == QMessageBox.Cancel:
            msgBox.close()   
    def close_window(self):
        global_var.state_machine = 1
        global_var.state_machine_flag = 1
        print('log', global_var.state_machine_flag, global_var.state_machine)

    def read_hrn_file_DB(self,event):
        self.hrn_table.clear()
        print("GV.stage1...",GV.stage)
        GV.stage=self.comboBox_2.currentIndex()+1
        print("GV.stage2....",GV.stage)
        DownloadHarnessData(GV.Location_No,GV.stage)


        GV.data_Available=4  
        
    def state_changed(self):
        self.pushButton_7.setStyleSheet("border-image: url(:/images/final_assets/Secondary_btn/exp_db.png);")
        self.hrn_table.clear()
  
            
    def HW_pin_convert(self,a):
        print("Hardware pin convertion")
        print(a)
        ckt_net=[]
        read=[]
        circuits=a
        try:
            for i in range(len(circuits)):
                read=(circuits[i])
                ckt=[]
                for j in range(len(read)):
                    if(read[j] in GV.virtual_pins):
                        ind=GV.virtual_pins.index(read[j])
                        ckt.append(GV.Actual_pins[ind])
            
                ckt_net.append(ckt)
               
        except(IndexError):
            print("No data  in circuit")
        return ckt_net



    def first_stage_learn(self,event):
        GV.Sflag=0
        GV.method = self.comboBox.currentIndex()
        
        GV.stage = self.comboBox_2.currentIndex() + 1
        if(GV.stage==1 or GV.stage==0):
            self.comboBox_2.clear()
            self.comboBox_2.addItem(str(1))
            

        if (GV.method==0):
            self.hrn_table.clear()
            GV.module_no=4
        elif(GV.method==1):
            self.hrn_table.clear()
            global_var.state_machine = 22
            global_var.state_machine_flag = 1
            print("learn using Cutting Chart")
        elif(GV.method==2):
            self.hrn_table.clear()
            self.usb_transfer()
            print("learn by USB")
            

        
            
    def edit_txt_data(self):
        self.hrn_table.setEnabled(True) 
        self.pushButton_4.setStyleSheet("border-image: url(:/images/final_assets/Secondary_btn/exp_db.png);")
       
    def usb_transfer(self):
        if(GV.stage==1):
            DeleteHarnessData(GV.Location_No)
        usb_detect = len(next(os.walk('/media/usb0'))[1])
        print("GV.Sflag at USB",GV.Sflag)
        if(usb_detect>0):
            file_name=QFileDialog.getOpenFileName(self, 'Open file', '/media/usb0')
            try:
                if(GV.stage==0):
                    GV.stage=1
                excel_file = openpyxl.load_workbook(file_name)
                Sheet_name = excel_file.sheetnames
                Circuits = excel_file[Sheet_name[GV.stage-1]]
                row = Circuits.max_row
                col = Circuits.max_column
                lrn_list = [[] for i in range(2048)]

                for i in range(1,row+1):
                    for j in range(1,col+1):
                        try:
                            item = Circuits.cell(i+1, j).value
                            try:
                                if(item==None):
                                    pass
                                else:
                                    lrn_list[i].append(int(item))
                                    # row += 1
                            except ValueError as e:
                                print('lrn_table valueerror', e)
                        except AttributeError:
                            row += 1
                try:            
                    read = []
                    for i in range(1, len(lrn_list)):
                        if (len(lrn_list[i]) > 0):
                            read.append(lrn_list[i])
                    print("Read",read)
                    # GV.circuits=read
                    GV.circuits=self.HW_pin_convert(read)
                    
                    GV.data_Available=4        
                except TypeError:
                    print("error while reading excel")            
            
            except FileNotFoundError:
                self.quit_msg="File Not Found "
                self.popupmeassage()
        else:
            self.quit_msg="USB Not Detected "
            self.popupmeassage()
    


    def save_hrn(self,event):
        self.hrn_table.setEnabled(False)
        print("Save Hrn")
        self.pushButton_6.setStyleSheet("background-image: url(:/images/final_assets/Main_Btn/btn_press.png);color: rgb(255, 255, 255);")
        if(GV.method==0):
##            circuits=GV.circuits
            row=self.hrn_table.rowCount()
            col=self.hrn_table.columnCount()
            print(row,col)
            lrn_list=[[] for i in range(2048)]

            for i in range(row):
                for j in range(col):
                    try:
                        item = self.hrn_table.item(i, j).text()
##                        print("lrn_list[i]",type(lrn_list[i]),lrn_list[i],type(item),item)
                        try:
                            lrn_list[i].append(int(item))
                            row+=1
##                            print("lrn_list[i]",lrn_list[i])
                        except ValueError as e:
                            print('lrn_table valueerror',e)
                    except AttributeError:
                        row+= 1
##            print('lrn_list',lrn_list)
            readcheck=[]
            for i in range(1,len(lrn_list)):
                if(len(lrn_list[i])>0):
                    readcheck.append(sorted(lrn_list[i]))
##            print('readcheck',readcheck)
            read=[]
            for x in range(len(readcheck)):
                readcheck[x]=[i for i in readcheck[x] if i != 0]
                if(len(readcheck[x])>0):
                    read.append(readcheck[x])
            
##            print('read',read)       
            a=sorted(read, key=operator.itemgetter(0))
            # print("GV.Actual_pins",GV.Actual_pins)
            # print("GV.virtual_pins",GV.virtual_pins)
            
            read_len=map(len,a)
            first_row_pt=sum(read_len)
            self.hrn_table.clear()
            GV.circuits=self.HW_pin_convert(a)
            
            self.hrn_table.setItem(0,0, QTableWidgetItem(str(first_row_pt)))
            
            for i in range(len(a)):
                b=a[i]
                for j in range(len(b)):
                    display_pts=b[j]
                    self.hrn_table.setItem(i+1,j, QTableWidgetItem(str(display_pts)))
            

            circuits=GV.circuits
            Wire_Type=[]
            Wire_color1=[]
            Wire_color2=[]
            Wire_Gauge=[]
            for i in range (len(circuits)):
                Type="cnt"
                value= 0
                Tolerence=0
                Wire_Type.append('W1')
                Wire_color1.append('WHT')
                Wire_color2.append('WHT')
                Wire_Gauge.append('0.5')

            UploadHarnessData(GV.Location_No,GV.stage,Type,value,Tolerence,circuits,Wire_Type,Wire_color1,Wire_color2,Wire_Gauge)
            UploadSysLog_Data(GV.Location_No,GV.Part_Name,'W/H learn And Stored',' Hrn Saved')
            Get_num_stages(GV.Location_No)
            self.comboBox_2.addItem(str(GV.stage+1))
            self.quit_msg="Data Saved "
            self.popupmeassage()
            print('db saved')
        if(GV.method==2):
            row=self.hrn_table.rowCount()
            col=self.hrn_table.columnCount()
            print(row,col)
            lrn_list=[[] for i in range(2048)]

            for i in range(row):
                for j in range(col):
                    try:
                        item = self.hrn_table.item(i, j).text()
##                        print("lrn_list[i]",type(lrn_list[i]),lrn_list[i],type(item),item)
                        try:
                            lrn_list[i].append(int(item))
                            row+=1
##                            print("lrn_list[i]",lrn_list[i])
                        except ValueError as e:
                            print('lrn_table valueerror',e)
                    except AttributeError:
                        row+= 1
##            print('lrn_list',lrn_list)
            readcheck=[]
            for i in range(1,len(lrn_list)):
                if(len(lrn_list[i])>0):
                    readcheck.append(sorted(lrn_list[i]))
##            print('readcheck',readcheck)
            read=[]
            for x in range(len(readcheck)):
                readcheck[x]=[i for i in readcheck[x] if i != 0]
                if(len(readcheck[x])>0):
                    read.append(readcheck[x])
            
##            print('read',read)       
            a=sorted(read, key=operator.itemgetter(0))
            # print("GV.Actual_pins",GV.Actual_pins)
            # print("GV.virtual_pins",GV.virtual_pins)
            GV.circuits=a
            read_len=map(len,GV.circuits)
            first_row_pt=sum(read_len)
            self.hrn_table.clear()
            print("GV.circuits........",GV.circuits)
            
            self.hrn_table.setItem(0,0, QTableWidgetItem(str(first_row_pt)))
            
            for i in range(len(a)):
                b=a[i]
                for j in range(len(b)):
                    display_pts=b[j]
                    self.hrn_table.setItem(i+1,j, QTableWidgetItem(str(display_pts)))
            GV.circuits=self.HW_pin_convert(a)
            circuits=GV.circuits
            Wire_Type=[]
            Wire_color1=[]
            Wire_color2=[]
            Wire_Gauge=[]
            for i in range (len(circuits)):
                Type="cnt"
                value= 0
                Tolerence=0
                Wire_Type.append('W1')
                Wire_color1.append('WHT')
                Wire_color2.append('WHT')
                Wire_Gauge.append('0.5')

            UploadHarnessData(GV.Location_No,GV.stage,Type,value,Tolerence,circuits,Wire_Type,Wire_color1,Wire_color2,Wire_Gauge)
            UploadSysLog_Data(GV.Location_No,GV.Part_Name,'W/H learn And Stored',' Hrn Saved')
            Get_num_stages(GV.Location_No)
            self.comboBox_2.addItem(str(GV.stage+1 ))
            self.quit_msg="Data Saved "
            self.popupmeassage()
            print('db saved')

    
    def import_Cutting_Chart(self,event):
        if (global_var.cutting_import==1):
            global_var.cutting_import=0

            self.pixmap = QPixmap('/home/pi/Desktop/AWHT/UI_files/final_assets/Checkbox/check.png')
            self.label_12.setPixmap(self.pixmap)
            self.label_12.clear()
            self.hrn_table.clear()
            print("global_var.cutting_import",global_var.cutting_import)

        elif (global_var.cutting_import==0):
            global_var.cutting_import=1
            
            
            print("global_var.cutting_import1s",global_var.cutting_import)

        global_var.state_machine=3
        global_var.state_machine_flag=1
        print(global_var.state_machine_flag,global_var.state_machine)
        

##def main():
##    app = QtGui.QApplication(sys.argv)
##    GUI = teaching_page3()  
##    GUI.show()
##    app.exec_()
##
##if __name__ == '__main__':
##    main()
##        
