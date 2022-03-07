import sys
sys.path.insert(1,'/home/pi/Desktop/AWHT/code/tester_files')
from Sql_db import *
from openpyxl import load_workbook
import global_var
import global_test_var as GV
from global_files import*

def clickMethod(mess):
        QtGui.QMessageBox.about("WARNING", mess)


def CuttingChart_import(From_pos,Cavity_no_pos,To_pos,Cavity_no1_pos,Name_pos,Color_pos,WireGauge_pos,Splice_name,path):
    print(path)
    ##########################################################################################################################################
    sheet=[]
    workbook = load_workbook(path)
    sheet= workbook.sheetnames
    worksheet = workbook[sheet[(GV.stage-1)]]
    ##########################################################################################################################################
    ## Ram variable
    From=[]
    Cavity_no=[]
    To=[]
    Cavity_no1=[]
    Name=[]
    Color=[]
    WireGauge=[]
    Splice_name=Splice_name
    Splice_max=0
    Splice_m=0
    splice_list=[]
    Splice_Replace=[]

    
    Name1=[]
    Color1=[]
    WireGauge1=[]
    ##########################################################################################################################################

    connlst = get_allConLib()

    #print(worksheet.max_row)

    for i in range(2,worksheet.max_row+1):
        From.append(str(worksheet.cell(i,From_pos).value))
        Cavity_no.append(str(worksheet.cell(i,Cavity_no_pos).value))
        To.append(str(worksheet.cell(i,To_pos).value))
        Cavity_no1.append(str(worksheet.cell(i,Cavity_no1_pos).value))
        Name.append(str(worksheet.cell(i,Name_pos).value))
        Color.append(str(worksheet.cell(i,Color_pos).value))
        WireGauge.append(str(worksheet.cell(i,WireGauge_pos).value))


    for i in range (len(From)):
        if(Splice_name not in From[i]):
            From[i]=From[i]+"-"+Cavity_no[i]
        if(Splice_name  not in To[i]):
            To[i]=To[i]+"-"+Cavity_no1[i]
##    print("From",From)
##    print("To  ",To)
##    print("Splice_name  ",Splice_name)

    ##########################################################################################################################################
    ## find  no of splice name eg:             ['S-1', 'S-A', 'S-B']

    for i in range(len(From)):
        if(Splice_name in From[i]):
            if(From[i] not in splice_list ):
                splice_list.append(From[i])

    for i in range(len(To)):
        if(Splice_name in To[i]):
            if(To[i] not in splice_list ):
                splice_list.append(To[i])
##    print("splice_list",splice_list)
    ##########################################################################################################################################
    ## find splice position and its replacement cavity no ['Con 1-3', 'CON5-1', 'CON5-1']  replace all splice with respective splice replacement variable
    ##

    flag1=0
    flag2=0
    for i in range(len(splice_list)):
        flag1=0
        flag2=0
        replace=None
        
        try:
            x=From.index(splice_list[i])
            flag1=1
        except ValueError:
            flag1=0
            
        try:
            y=To.index(splice_list[i])
            flag2=1
        except ValueError:
            flag2=0
        if(flag1==1 and flag2==1):
            if(x>y):
                Splice_Replace.append(From[y])
                replace=From[y]
            else:
                Splice_Replace.append(To[x])
                replace=To[x]
        if(flag1==1 and flag2==0):
            Splice_Replace.append(To[x])
            replace=To[x]
        if(flag1==0 and flag2==1):
            Splice_Replace.append(From[y])
            replace=From[y]
            
        if(flag1==1 or flag2==1):
            for j in range(len(To)):
                if(splice_list[i]==To[j]):
                    To[j]=replace
                if(splice_list[i]==From[j]):
                    From[j]=replace
    ###########################################################################################################################################
    ## Remove same net eg: CON5-1 CON5-1 ALSO DELETE UNNESARY BLOCK FROM Name,COLOUR,WIREGAUGE
    GV.From_Temp=[]
    GV.To_Temp=[]
    GV.Color_Temp=[]
    GV.Name_Temp=[]
    GV.Color_Temp=[]
##    GV.WireGauge_Temp=[]
    for i in range(len(To)):
            if(From[i]!=To[i]):
                GV.To_Temp.append(To[i])
                GV.From_Temp.append(From[i])
                GV.Name_Temp.append(Name[i])
                GV.Color_Temp.append(Color[i])
                GV.WireGauge_Temp.append(WireGauge[i])


        
    print("From...",GV.From_Temp)
    print("To. .. ",GV.To_Temp)
    print("Name..",GV.Name_Temp)
    print("Color..",GV.Color_Temp)
    print("WireGauge..",GV.WireGauge_Temp)
##    x=cutting_chart_window()
##    x.loaddata(x)
    Type = 'Continuity'
    UploadCutting_ChartData(GV.Location_No,Type,GV.Name_Temp,GV.From_Temp,GV.To_Temp,GV.Color_Temp)
    
##########################################################################################################################################
    ## IMPORT DATA FROM SQDB TO RAM VARIABLE.
    HN1 = []
    HN2 = []
##    print(GV.From_Temp)
##    print(GV.To_Temp)
    print("GV.library_mapping",GV.library_mapping)
    for i in range(len(GV.From_Temp)):
        
        f = GV.From_Temp[i]
        t = GV.To_Temp[i]
    #    if not('None' in f):
        (conn1,q1) = f.split('-')
        print("conn1",conn1)
        
##        out1 = get_LibMap(conn1)[0][0]
        (conn2,q2) = t.split('-')
        
        print("conn2",conn2)
##        out2 = get_LibMap(conn2)[0][0]
        for i in range (len(GV.library_mapping)):
                if conn1 in GV.library_mapping[i]:
                        out1=GV.library_mapping[i][0]
                if conn2 in GV.library_mapping[i]:
                        out2=GV.library_mapping[i][0]
                        
        pt1 = get_Points(out1)
        pt2 = get_Points(out2)
##        print("out1",out1,out2)
##        print("pt1",pt1,q1)
##        print("pt2",pt2,q2)
    
        try:
             HN1.append(pt1[int(q1)-1][0])
             HN2.append(pt2[int(q2)-1][0])
        except:
            print('No data in Globalgrp2')
##    print('HN1',HN1)
##    print('HN2',HN2)
##########################################################################################################################################
##combine HN1 AND HN2
    ckt_lst=[[] for i in range(len(HN1))]
    ckt_lst1=[[] for i in range(len(HN1))]
    
    for i in range(len(HN1)):
        ckt_lst[i].append(HN1[i])
        ckt_lst[i].append(HN2[i])

##    print(ckt_lst)

    key=0
    maxnum=max(HN1)
    key=0
    remove_list=[]
    for i in range(maxnum):
        key=i+1
        for j in range(len(HN1)):
            if (key==HN1[j]):
##                print("i",i,j)
                if HN1[j] not in ckt_lst1[j]:
                    ckt_lst1[j].append(key)
                    ckt_lst1[j].append(HN2[j])
                    Color1.append(GV.Color_Temp[j])
                    Name1.append(GV.Name_Temp[j])
                    WireGauge1.append(GV.WireGauge_Temp[j])
                else:
                    ckt_lst1[i].append(HN2[j])
                    
    ckt_pt=[]
    cktpionts=[]
    Wire_Type=[]
    Wire_color1=[]
    Wire_color2=[]
    Wire_Gauge=[]
    for i in ckt_lst1:
        if (len(i))>0:
            ckt_pt.append(i)
            
    
    #print("ckt_pt",ckt_pt)
    for i in range(len(ckt_pt)):
        cktpionts.append(sorted(ckt_pt[i]))
    
##    Location_No = 1
    cktpionts=sorted(cktpionts)
    GV.Cutting_circuits = cktpionts
    
    list2 = [] 
    for val in GV.Cutting_circuits: 
        if not val[0] in list2: 
            list2.append(val[0]) 
    list2 = [[x] for x in list2] 
         
    for val in GV.Cutting_circuits: 
        for i in list2: 
            if val[0] == i[0]: 
                i.extend(val[1:])
    #print(list2)


    GV.Cutting_circuits=list2
    print("GV.Cutting_circuits",GV.Cutting_circuits)
    for i in range(len(GV.Cutting_circuits)):
             for j in range(len(ckt_pt)):
                     if GV.Cutting_circuits[i][0] in ckt_pt[j]:
                             Wire_Type.append(GV.Name_Temp[ckt_pt.index(ckt_pt[j])])
                             Wire_color1.append(GV.Color_Temp[ckt_pt.index(ckt_pt[j])])
                             Wire_color2.append(GV.Color_Temp[ckt_pt.index(ckt_pt[j])])
                             Wire_Gauge.append(GV.WireGauge_Temp[ckt_pt.index(ckt_pt[j])])
                             
                             break
        
    Type="cnt"
    value= 0
    Tolerence=0
    print("GV.stage",GV.stage)
##    UploadHarnessData(GV.Location_No,GV.Cutting_circuits,Wire_Type,Wire_color1,Wire_color2,Wire_Gauge)
    UploadHarnessData(GV.Location_No,GV.stage,Type,value,Tolerence,GV.Cutting_circuits,Wire_Type,Wire_color1,Wire_color2,Wire_Gauge)
    

##########################################################################################################################################
##if __name__ == '__main__':
##    ## following Argument needed
##    From_pos=1
##    Cavity_no_pos=2
##    To_pos=3
##    Cavity_no1_pos=4
##    Name_pos=5
##    Color_pos=6
##    WireGauge_pos=7
##    Splice_name="sp-"
##    path="/home/pi/Desktop/CuttingChartData.xlsx"
##    CuttingChart_import(From_pos,Cavity_no_pos,To_pos,Cavity_no1_pos,Name_pos,Color_pos,WireGauge_pos,Splice_name,path)

