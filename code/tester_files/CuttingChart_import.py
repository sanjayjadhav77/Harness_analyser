import sys
sys.path.insert(1,'/home/pi/Desktop/AWHT/code/tester_files')
from Sql_db import *
from openpyxl import load_workbook



def CuttingChart_import(From_pos,Cavity_no_pos,To_pos,Cavity_no1_pos,Name_pos,Color_pos,WireGauge_pos,Splice_name,path):
    ##########################################################################################################################################
    workbook = load_workbook('CuttingChartData.xlsx')
    worksheet = workbook['Cuttting_Chart']
    ##########################################################################################################################################
    ## Ram variable
    From=[]
    Cavity_no=[]
    To=[]
    Cavity_no1=[]
    Name=[]
    Color=[]
    WireGauge=[]
    Splice_name="S-"
    Splice_max=0
    Splice_m=0
    splice_list=[]
    Splice_Replace=[]

    From_Temp=[]
    To_Temp=[]
    Name_Temp=[]
    Color_Temp=[]
    WireGauge_Temp=[]

    Name1=[]
    Color1=[]
    WireGauge1=[]
    ##########################################################################################################################################

    connlst = get_allConLib()

    #print(worksheet.max_row)

    for i in range(2,worksheet.max_row+1):
        From.append(str(worksheet.cell(i,From_pos).value))
        Cavity_no.append(str(worksheet.cell(i,Cavity_no_pos).value))
        To.append(str(worksheet.cell(i,To_pos).value))
        Cavity_no1.append(str(worksheet.cell(i,Cavity_no1_pos).value))
        Name.append(str(worksheet.cell(i,Name_pos).value))
        Color.append(str(worksheet.cell(i,Color_pos).value))
        WireGauge.append(str(worksheet.cell(i,WireGauge_pos).value))

    ##########################################################################################################################################
    ## Join  connector name and cavity no eg-Con1-1 con2-3 from both TO and From table
    ##  From ['Con 1-1', 'Con 1-2', 'Con 1-2', 'Con 1-3', 'S-1', 'S-1', 'S-A', 'S-A', 'S-A', 'S-B']
    ##  To   ['Con 3-1', 'Con 3-2', 'Con 3-3', 'S-1', 'Con 3-4', 'Con 4-1', 'CON5-1', 'CON7-1', 'S-B', 'CON6-1']

    for i in range (len(From)):
        if(Splice_name not in From[i]):
            From[i]=From[i]+"-"+Cavity_no[i]
        if(Splice_name  not in To[i]):##if __name__ == '__main__':

            To[i]=To[i]+"-"+Cavity_no1[i]
    ##print("From",From)
    ##print("To  ",To)

    ##########################################################################################################################################
    ## find  no of splice name eg:             ['S-1', 'S-A', 'S-B']

    for i in range(len(From)):
        if(Splice_name in From[i]):
            if(From[i] not in splice_list ):
                splice_list.append(From[i])

    for i in range(len(To)):
        if(Splice_name in To[i]):
            if(To[i] not in splice_list ):
                splice_list.append(To[i])
    #print(splice_list)
    ##########################################################################################################################################
    ## find splice position and its replacement cavity no ['Con 1-3', 'CON5-1', 'CON5-1']  replace all splice with respective splice replacement variable
    ##

    flag1=0
    flag2=0
    for i in range(len(splice_list)):
        flag1=0
        flag2=0
        replace=None
        
        try:
            x=From.index(splice_list[i])
            flag1=1
        except ValueError:
            flag1=0
            
        try:
            y=To.index(splice_list[i])
            flag2=1
        except ValueError:
            flag2=0
        if(flag1==1 and flag2==1):
            if(x>y):
                Splice_Replace.append(From[y])
                replace=From[y]
            else:
                Splice_Replace.append(To[x])
                replace=To[x]
        if(flag1==1 and flag2==0):
            Splice_Replace.append(To[x])
            replace=To[x]
        if(flag1==0 and flag2==1):
            Splice_Replace.append(From[y])
            replace=From[y]
            
        if(flag1==1 or flag2==1):
            for j in range(len(To)):
                if(splice_list[i]==To[j]):
                    To[j]=replace
                if(splice_list[i]==From[j]):
                    From[j]=replace
    ##########################################################################################################################################
    ## Remove same net eg: CON5-1 CON5-1 ALSO DELETE UNNESARY BLOCK FROM Name,COLOUR,WIREGAUGE

    for i in range(len(To)):
            if(From[i]!=To[i]):
                To_Temp.append(To[i])
                From_Temp.append(From[i])
                Name_Temp.append(Name[i])
                Color_Temp.append(Color[i])
                WireGauge_Temp.append(WireGauge[i])


##    print("From",From_Temp)
##    print("To  ",To_Temp)
##    print("Name",Name_Temp)
##    print("Color",Color_Temp)
##    print("WireGauge",WireGauge_Temp)
##########################################################################################################################################
    ## IMPORT DATA FROM SQDB TO RAM VARIABLE.
    HN1 = []
    HN2 = []
    for i in range(len(From_Temp)):
        f = From_Temp[i]
        t = To_Temp[i]
    #    if not('None' in f):
        (conn1,q1) = f.split('-')
        out1 = get_LibMap(conn1)[0][1]
        out2 = get_LibMap(conn2)[0][1]
        pt1 = get_Points(out1)
        pt2 = get_Points(out2)
    #    print(conn2)
        try:
             HN1.append(pt1[int(q1)-1][0])
             HN2.append(pt2[int(q2)-1][0])
        except:
            print('error')
##    print('HN1',HN1)
##    print('HN2',HN2)
##########################################################################################################################################
##combine HN1 AND HN2
    ckt_lst=[[] for i in range(len(HN1))]
    ckt_lst1=[[] for i in range(len(HN1))]
    
    for i in range(len(HN1)):
        ckt_lst[i].append(HN1[i])
        ckt_lst[i].append(HN2[i])

##    print(ckt_lst)

    key=0
    maxnum=max(HN1)
    key=0
    remove_list=[]
    for i in range(maxnum):
        key=i+1
        for j in range(len(HN1)):
            if (key==HN1[j]):
                if HN1[j] not in ckt_lst1[i]:
                    ckt_lst1[i].append(key)
                    ckt_lst1[i].append(HN2[j])
                    Color1.append(Color_Temp[j])
                    Name1.append(Name_Temp[j])
                    WireGauge1.append(WireGauge_Temp[j])
                else:
                    ckt_lst1[i].append(HN2[j])
                    
    ckt_pt=[]
    for i in ckt_lst1:
        if (len(i))>0:
            ckt_pt.append(i)
            
    
##    print("ckt_pt",ckt_pt)

    Location_No = 1
    circuits = ckt_pt
    Wire_Type = Name1
    Wire_color1 = Color1
    Wire_color2 = Color1
    Wire_Gauge = WireGauge1

    print("circuits",circuits)
    print("Wire_Type",Wire_Type)
    print("Wire_color1",Wire_color1)
    print("Wire_color2",Wire_color2)
    print("WireGauge",Wire_Gauge)

   
    UploadHarnessData(GV.Location_No,circuits,Wire_Type,Wire_color1,Wire_color2,Wire_Gauge)

##########################################################################################################################################
if __name__ == '__main__':
    ## following Argument needed
    From_pos=1
    Cavity_no_pos=2
    To_pos=3
    Cavity_no1_pos=4
    Name_pos=5
    Color_pos=6
    WireGauge_pos=7
    Splice_name="S-"
    path="x"
    CuttingChart_import(From_pos,Cavity_no_pos,To_pos,Cavity_no1_pos,Name_pos,Color_pos,WireGauge_pos,Splice_name,path)
