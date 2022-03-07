import global_test_var as GV
from global_files import*
import time
import global_var
from  Sql_db import *
from weekno import*
import fo_main 
import time, pprint, cups
import datetime
from datetime import datetime,date
import pandas as pd
from openpyxl.styles import Alignment,Font ,Border,Side,Protection 
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from printer import*
import subprocess
global hold
hold=0
##################################################################################

def lbl_printing(print_data):
##    Serial_Init()
##    Serial_Printing()
    print_data=str(print_data)
    f= open("/home/pi/Desktop/.HA_Editor/Documents/1.LBL","w")
    f.write(str(print_data))
    f.close()

    currentDate=datetime.now()
   
    
    GV.HMS=currentDate.strftime("%H:%M:%S")
    GV.hrs=currentDate.strftime("%H")
    GV.mint=currentDate.strftime("%M")
    GV.sec=currentDate.strftime("%S")
    GV.DMY=currentDate.strftime("%d/%m/%Y")
    GV.date=currentDate.strftime("%d")
    GV.mon=currentDate.strftime("%m")
    GV.year=currentDate.strftime("%Y")
   

    localtime = time.localtime(time.time())
    my_calendar = CustomizedCalendar(start_weekday=WEEKDAY.SUN)
    week_no = my_calendar.calculate(datetime(localtime[0],localtime[1],localtime[2]))[1]
    ##print(my_calendar.calculate(datetime(localtime[0],localtime[1],localtime[2]))[1])    
    # print("week_no",week_no)
    if(week_no<=9):
        GV.week_no = '0'+ str(week_no)
        # print(week_number,"week_number")
    else:
        GV.week_no = str(week_no)

    if(int(GV.Pass_Count)<=9):
        PC='0'+'0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=99):
        PC='0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999):
        PC='0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=9999):
        PC='0'+'0'+str(GV.Pass_Count)    
    elif(int(GV.Pass_Count)<=99999):
        PC='0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999999):
        PC=str(GV.Pass_Count)
    else:
        PC=GV.Pass_Count
    #daycount-------------------------------------------------------------------------------
    date=GV.DMY
    d = list(map(int,date.split("/")))
    days = [0,31,28,31,30,31,30,31,31,30,31,30,31]
    if d[2] % 400 == 0:
        days[2]+=1
    elif d[2]%4 == 0 and d[2]%100!=0:
        days[2]+=1
    for i in range(1,len(days)):
        days[i]+=days[i-1]
    GV.daycount=str(days[d[1]-1]+d[0])
    #week day-----------------------------------------------------------------------
    
    if(int(localtime[6]==6)):#for sunday localtime 6=6....so to start week from sunday minus 5 is added
        GV.week_day = (str(localtime[6]-5))
    else:
        GV.week_day = (str(localtime[6]+2)) #from monday localtime 6 = 0 (to 5)so plus 2 is added
        ##print(Week_day,"Week_day")
    #week no-------------------------------------------------------------------------------------------------
    if((GV.HMS>='06:00:00')&(GV.HMS<='13:59:59')): # 06AM to 02PM(shift A)
        GV.shift = 'A'
    elif((GV.HMS>='14:00:00')&(GV.HMS<='21:59:59')): # 02PM to 10PM(shift B)
        GV.shift = 'B'
    elif((GV.HMS>='22:00:00')&(GV.HMS<='23:59:59')): # 10PM to 12AM(shift C)
        GV.shift = 'C'
    elif((GV.HMS>='00:00:00')&(GV.HMS<='05:59:59')): # 12AM to 06AM(shift C)
        GV.shift = 'C'
    else:
        print('no shift')
    
    year=GV.year[2:]
    # print('GV.year',year)
    
    #read_data= open("/home/pi/Desktop/Harness-lbl.LBL","r")
    read_data= open("/home/pi/Desktop/.HA_Editor/Documents/1.LBL","r")
    file_read=read_data.read()
    read_data.close()

    file_read=file_read.replace("@1",(GV.DMY))
    file_read=file_read.replace("@2",(GV.HMS))
    file_read=file_read.replace("@4",(GV.mon))
    file_read=file_read.replace("@5",(year))
    file_read=file_read.replace("@6",str(PC))
    file_read=file_read.replace("@8",(GV.date))
    file_read=file_read.replace("@C",str(GV.code_1))
    file_read=file_read.replace("@B",str(GV.code_2))
    file_read=file_read.replace("@X",(GV.week_day))
    file_read=file_read.replace("@Y",(GV.daycount))
    file_read=file_read.replace("@N",(GV.week_no))
    file_read=file_read.replace("@D",(GV.shift))
    file_read=file_read.replace("@H",(GV.hrs))
    file_read=file_read.replace("@M",(GV.mint))
    file_read=file_read.replace("@S",(GV.sec))
    # print("print data-",file_read)

    f= open("/home/pi/Desktop/.HA_Editor/Documents/1.LBL","w")
    f.write(str(file_read))
    f.close()

    conn = cups.Connection()
    printers = conn.getPrinters ()
    pprint.pprint(printers)
 
    printer = conn.getDefault()
    # print("Default1:", printer)
     
    if printer == None:
        printer = list(printers.keys())[0]
        # print("Default2:", printer)
     
    myfile = '/home/pi/Desktop/.HA_Editor/Documents/1.LBL'
    pid = conn.printFile(printer, myfile, "test", {})
    status=subprocess.getoutput("lpstat -t")
    # print(status)
    string="Waiting for printer to become available."
    if(string in status):
        print("printer not available")
    else:    
        print("label printed sucessfully")


        
################################################################################            
def bar1_file_replace():
    bar_data=str(GV.Local_Barcode1_Data)
    GV.bar1_data=str(bar_data)
    print('bar1_data',GV.bar1_data)
    if(int(GV.Pass_Count)<=9):
        PC='0'+'0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=99):
        PC='0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999):
        PC='0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=9999):
        PC='0'+'0'+str(GV.Pass_Count)    
    elif(int(GV.Pass_Count)<=99999):
        PC='0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999999):
        PC=str(GV.Pass_Count)
    else:
        PC=GV.Pass_Count
     
    GV.week_day='0'+str(GV.week_day)
    
    GV.bar1_data=GV.bar1_data.replace("@1",(GV.DMY))
    GV.bar1_data=GV.bar1_data.replace("@2",(GV.HMS))
    GV.bar1_data=GV.bar1_data.replace("@4",(GV.mon))
    GV.bar1_data=GV.bar1_data.replace("@5",(GV.year))
    GV.bar1_data=GV.bar1_data.replace("@6",str(PC))
    GV.bar1_data=GV.bar1_data.replace("@8",(GV.date))
    GV.bar1_data=GV.bar1_data.replace("@X",(GV.week_no))
    GV.bar1_data=GV.bar1_data.replace("@C",str(GV.code_1))
    GV.bar1_data=GV.bar1_data.replace("@B",str(GV.code_2))
    GV.bar1_data=GV.bar1_data.replace("@N",(GV.week_day))
    
    
def bar2_file_replace():
    bar_data=str(GV.Local_Barcode2_Data)
    GV.bar2_data=str(bar_data)

    if(int(GV.Pass_Count)<=9):
        PC='0'+'0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=99):
        PC='0'+'0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999):
        PC='0'+'0'+'0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=9999):
        PC='0'+'0'+str(GV.Pass_Count)    
    elif(int(GV.Pass_Count)<=99999):
        PC='0'+str(GV.Pass_Count)
    elif(int(GV.Pass_Count)<=999999):
        PC=str(GV.Pass_Count)
    else:
        PC=GV.Pass_Count
    GV.week_day='0'+str(GV.week_day)
    
    GV.bar2_data=GV.bar2_data.replace("@1",(GV.DMY))
    GV.bar2_data=GV.bar2_data.replace("@2",(GV.HMS))
    GV.bar2_data=GV.bar2_data.replace("@4",(GV.mon))
    GV.bar2_data=GV.bar2_data.replace("@5",(GV.year))
    GV.bar2_data=GV.bar2_data.replace("@6",str(PC))
    GV.bar2_data=GV.bar2_data.replace("@8",(GV.date))
    GV.bar2_data=GV.bar2_data.replace("@X",(GV.week_no))
    GV.bar2_data=GV.bar2_data.replace("@C",str(GV.code_1))
    GV.bar2_data=GV.bar2_data.replace("@B",str(GV.code_2))
    GV.bar2_data=GV.bar2_data.replace("@N",(GV.week_day))
    
def barcode1_scaning(self):   ##barcode scan
    # print("Brcode scaning")
    GV.Visual_Engine_Start=7
    self.p4.barcode_scan_line.setFocus()  
        # GV.data_delivered=16   
    if(len(GV.scan_data)>0):
        bar1_file_replace()
        self.p4.barcode_scan_line.clear()                            
        if(GV.scan_data==(GV.bar1_data)):
            GV.scan_data=''
            print("Bar1 match")
            self.p4.msg_line.setText("Bar1 match")
            self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
            time.sleep(0.2)
            self.p4.barcode_scan_line.clear()
##            self.p4.barcode_scan_line.clearFocus()
            GV.Hold=0
            GV.Holdstate=10
            GV.Estate=14
##            GV.data_delivered=0
                

        else:
            GV.scan_data=''
            print("Bar1 not match")
            self.p4.msg_line.setText("Bar1 not match")
            self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
            # time.sleep(0.2)
            self.p4.barcode_scan_line.clear()
            GV.Hold=0
            GV.Holdstate=9
            GV.Estate=14
            # GV.data_delivered=15      
                    
def barcode2_scaning(self):   ##barcode scan
    self.p4.barcode_scan_line.setFocus()
    if(len(GV.scan_data)>0):
        self.p4.barcode_scan_line.clear()
        bar2_file_replace()
        print('bar222',GV.bar2_data)
        if(GV.scan_data==(GV.bar2_data)):
            GV.scan_data=''
            print("Bar2 match")
            self.p4.msg_line.setText("Bar2 match")
            self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
            
            self.p4.barcode_scan_line.clear()
##            self.p4.barcode_scan_line.clearFocus()
##            GV.data_delivered=0
            GV.Hold=0
            GV.Holdstate=11
            GV.Estate=14
        else:
            GV.scan_data=''
            print("Bar2 not match")
            self.p4.msg_line.setText("Bar2 not match")
            self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
            # time.sleep(0.2)
            self.p4.barcode_scan_line.clear()
            GV.Hold=0
            GV.Holdstate=10
            GV.Estate=14   
def Autopartload(self):
    
    if(len(GV.scan_data)>0):
        part_name = GV.scan_data[int(GV.PartNO_Length):(int(GV.PartNO_Length)+len(GV.PartNoLoc))]
        print("part_name",part_name,GV.PartNO_Length)
        # print("GV.PartNoLoc",len(GV.PartNoLoc))
        sr_no = GV.scan_data[int(GV.Var1_Length):(int(GV.Var1_Length)+len(GV.UserVar1_Loc))]
        print("sr_no",sr_no)
        GV.scan_data=''
        for i in range (len(GV.part_name_list)):
            if(part_name in GV.part_name_list[i]):
                Location_No=GV.part_name_list[i][0]
                P_Name=GV.part_name_list[i][1]
                if(GV.Location_No!=Location_No):
                    GV.Location_No=Location_No
                    GV.Part_Name=P_Name
                    global_var.cable_change_flag=1
                    GV.Local_Cable_Info=DownloadCable_Info(GV.Location_No)
                    GV.Pass_Count=GV.Local_Cable_Info[0][0]
                    GV.Fail_Count=GV.Local_Cable_Info[0][1]
                    print("GV.fail_Count",GV.Pass_Count,GV.Fail_Count,GV.Part_Name,GV.Location_No)
                   
                    self.pass_cnt_line.setText(str(GV.Pass_Count))
                    self.fail_cnt_line.setText(str(GV.Fail_Count))
                    self.fo_qty.setText(str(GV.FoQty))
                    self.label_19.setText(str(GV.Part_Name))
                    self.label_20.setText(str(GV.Location_No))

                    self.barcode_scan_line.clear()
                    GV.scan_complete=1
                    self.msg_line.setText("Press Enter...")
                    self.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")

                    
                else:
                    GV.scan_complete=1
##                    self.barcode_scan_line.clearFocus()
                    self.msg_line.setText("Press Enter...")
                    self.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")

                    print("select Another Cable")
                
                
            else:
                self.barcode_scan_line.clear()
        self.barcode_scan_line.clear()
    else:
        self.barcode_scan_line.clear()

def Fixtures_scaning(self):   ##Asscet scan
##    print("GV.scan_data",GV.scan_data)
    self.p4.barcode_scan_line.setFocus()
    self.p4.msg_line.setText(GV.prod_data[GV.barcount][2])
    self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
    if(len(GV.scan_data)>0):
        #self.p4.barcode_scan_line.clear()                            
        if(GV.scan_data==(str(GV.prod_data[GV.barcount][1]))):
            print("Asset Barcode matched")
            self.p4.barcode_scan_line.clear()
            
            GV.barcount = GV.barcount+1
            self.p4.msg_line.setText("Asset Barcode matched")
            time.sleep(0.4)
            GV.scan_data=''
            if(GV.barcount>(int(GV.Noofbar)-1)):
                GV.Estate=15
                GV.data_delivered=4
                GV.barcount=0
                print("bar scan complet")

                GV.Set_fo_flag=1
##                self.p4.barcode_scan_line.clearFocus()
                
                if (GV.Set_fo_flag==1):
                    GV.Set_fo_flag=0
                    self.new_window =fo_main.Ui_Dialog()
                    self.new_window.show()
                        
        else:
            GV.scan_data=''
            print("Asset Barcode not matched")
            
            self.p4.msg_line.setText("Asset Barcode not matched")
            time.sleep(0.6)
            self.p4.barcode_scan_line.clear()
def Report_Generation(self):
    
    GV.SERIAL_NUMBER='7'+'_pass'
    storage_path = '/home/pi/Desktop/.HA_Editor/Reports'+ '/' + str(GV.Location_No)
    try:
        os.makedirs(storage_path)
    except FileExistsError:
        print('folder exist')
        
    file_path = storage_path  + '/' + str(GV.SERIAL_NUMBER) + '.xlsx'
    self.p4.msg_line.setText("Generating Report...")
    self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
    GV.cutting_chartData=DownloadCutting_ChartData(GV.Location_No)
##    print("GV.cutting_chartData....",GV.cutting_chartData)
    localtime = time.localtime(time.time())
    time_current = (str(localtime[3])+':'+str(localtime[4])+':'+str(localtime[4]))
    today = date.today()
    date_now = today.strftime("%b-%d-%Y")
    Oprator_Name=GV.code_1
    temp_data=GV.cutting_chartData
##    print("temp_data............",temp_data)
    for i in range(len(temp_data)):
            value = 'Pass'
            temp_data[i]=temp_data[i] + (value,)

    df = pd.DataFrame(temp_data, columns =['Type','Name', 'From','To', 'Color', 'Status'])
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name='Sheet1', index=False,startrow=5)
    writer.save()
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb['Sheet1']
    logo = Image("/home/pi/Desktop/HA_Editor/UI_files/final_assets/KalpLogo.bmp")
    logo.height = 40
    logo.width = 40
    ws.row_dimensions[1].height = 30
    ws.add_image(logo, "A1")
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.merge_cells('B1:E1')
    ws.merge_cells('A2:B2')
    ws.merge_cells('A3:B3')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A5:B5')
    ws.merge_cells('C2:D2')
    ws.merge_cells('C3:D3')
    ws.merge_cells('C4:D4')
    ws.merge_cells('C5:D5')
    ws.cell(row=1, column=2, value="KalpTech Solutions PVT LTD ")
    cell = ws.cell(row=1, column=2)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    stp=GV.SERIAL_NUMBER.split('_')
    GV.SERIAL_NUMBER=stp[0]
    ws.cell(row=2, column=1, value="Serial Number :")
    ws.cell(row=3, column=1, value="Date :")
    ws.cell(row=4, column=1, value="Time :")
    ws.cell(row=5, column=1, value="Operator Name :")
    ws.cell(row=2, column=3, value=GV.SERIAL_NUMBER)
    ws.cell(row=3, column=3, value=date_now)
    ws.cell(row=4, column=3, value=time_current)
    ws.cell(row=5, column=3, value=Oprator_Name)



    wb.save(file_path)

    os.system('unoconv -f pdf '+file_path+'')
    print('Done',file_path)
    self.p4.msg_line.setText("Report Complete...")
    self.p4.msg_line.setStyleSheet("""border-radius: 1px;font: 20pt "Roboto [GOOG]";color: blue""")
    time.sleep(0.6)
    GV.Estate=12

#----------------------------------Thread-------------------------------------------------------------------------------------------#
######################################################################################################################################################

##if __name__ == '__main__':
##    print("barcode scanning")
 
    
    
